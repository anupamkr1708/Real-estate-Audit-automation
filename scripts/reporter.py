import argparse
import logging
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from typing import List, Dict   # <<< FIX ADDED


logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("reporter")


class ImprovedAuditReporter:
    """Enhanced reporter with detailed escalation schedules"""
    
    def __init__(self, output_filename: str):
        self.output_filename = output_filename
        self.wb = Workbook()
        
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        
        # Styles
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.bold_font = Font(bold=True)
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def create_section_1_lease_status(self, findings: list):
        """Section 1: Lease Status"""
        
        logger.info("Creating Section 1: Lease Status...")
        
        ws = self.wb.create_sheet("Audit Report")
        
        ws.merge_cells('A1:G1')
        ws['A1'] = "Data cells highlighted in Yellow will be updated in AE."
        
        row = 3
        ws[f'A{row}'] = "1. For the following Suites, there are discrepancies in Lease Status."
        ws[f'A{row}'].font = self.bold_font
        row += 2
        
        headers = ['Suite No.', 'Prior NKF', "Owner's Model", 'Rent Roll', 'Stacking Plan', 'Leases']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header)
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        row += 1
        
        if not findings:
            ws.cell(row, 1, "No discrepancies found")
        else:
            for finding in findings:
                ws.cell(row, 1, finding['suite_number']).border = self.border
                ws.cell(row, 2, finding['prior_nkf']).border = self.border
                ws.cell(row, 3, finding['owners_model']).border = self.border
                ws.cell(row, 4, finding['rent_roll']).border = self.border
                ws.cell(row, 5, finding['stacking_plan']).border = self.border
                ws.cell(row, 6, finding['leases']).border = self.border
                
                # Highlight mismatches
                if finding['prior_nkf'] != finding['owners_model']:
                    ws.cell(row, 2).fill = self.yellow_fill
                    ws.cell(row, 3).fill = self.yellow_fill
                
                row += 1
        
        return row + 2
    
    def create_section_2_tenant_name(self, findings: list, start_row: int):
        """Section 2: Tenant Names"""
        
        logger.info("Creating Section 2: Tenant Names...")
        
        ws = self.wb["Audit Report"]
        row = start_row
        
        ws[f'A{row}'] = "2. For the following Suite, there is discrepancy in Tenant Name."
        ws[f'A{row}'].font = self.bold_font
        row += 2
        
        headers = ['Suite No.', 'Prior NKF', "Owner's Model", 'Rent Roll', 'Stacking Plan', 'Leases']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header)
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        row += 1
        
        if not findings:
            ws.cell(row, 1, "No discrepancies found")
        else:
            for finding in findings:
                ws.cell(row, 1, finding['suite_number']).border = self.border
                ws.cell(row, 2, finding['prior_nkf']).border = self.border
                ws.cell(row, 3, finding['owners_model']).border = self.border
                ws.cell(row, 4, finding['rent_roll']).border = self.border
                ws.cell(row, 5, finding['stacking_plan']).border = self.border
                ws.cell(row, 6, finding['leases']).border = self.border
                
                # Highlight all names (mismatch)
                for col in [2, 3, 4, 5]:
                    if ws.cell(row, col).value != 'NAV':
                        ws.cell(row, col).fill = self.yellow_fill
                
                row += 1
        
        return row + 2
    
    def create_section_3_lease_expiration(self, findings: list, start_row: int):
        """Section 3: Lease Expiration"""
        
        logger.info("Creating Section 3: Lease Expiration...")
        
        ws = self.wb["Audit Report"]
        row = start_row
        
        ws[f'A{row}'] = "3. For the following Tenants, there are discrepancies in Lease Expiration Date (LED):"
        ws[f'A{row}'].font = self.bold_font
        row += 2
        
        headers = ['Suite No.', 'Tenant Name', 'Prior NKF', "Owner's Model", 'Rent Roll', 'Stacking Plan', 'Leases']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header)
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        row += 1
        
        if not findings:
            ws.cell(row, 1, "No discrepancies found")
        else:
            for finding in findings:
                ws.cell(row, 1, finding['suite_number']).border = self.border
                ws.cell(row, 2, finding['tenant_name']).border = self.border
                ws.cell(row, 3, finding['prior_nkf']).border = self.border
                ws.cell(row, 4, finding['owners_model']).border = self.border
                ws.cell(row, 5, finding['rent_roll']).border = self.border
                ws.cell(row, 6, finding['stacking_plan']).border = self.border
                ws.cell(row, 7, finding['leases']).border = self.border
                
                # Highlight dates
                for col in [3, 4, 5, 6]:
                    if ws.cell(row, col).value != 'NAV':
                        ws.cell(row, col).fill = self.yellow_fill
                
                row += 1
        
        return row + 2
    
    def create_section_4_leased_area(self, findings: list, start_row: int):
        """Section 4: Leased Area"""
        
        logger.info("Creating Section 4: Leased Area...")
        
        ws = self.wb["Audit Report"]
        row = start_row
        
        ws[f'A{row}'] = "4. For the following Tenants, there are discrepancies in Leased Area (SF):"
        ws[f'A{row}'].font = self.bold_font
        row += 2
        
        headers = ['Suite No.', 'Tenant Name', 'Prior NKF (A)', "Owner's Model (B)", 
                  'Rent Roll (C)', 'Stacking Plan (D)', 'Variance (A-B)', 'Variance (A-C)', 'Variance (A-D)']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header)
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        row += 1
        
        if not findings:
            ws.cell(row, 1, "No discrepancies found")
        else:
            for finding in findings:
                ws.cell(row, 1, finding['suite_number']).border = self.border
                ws.cell(row, 2, finding['tenant_name']).border = self.border
                ws.cell(row, 3, finding['prior_nkf_a']).border = self.border
                ws.cell(row, 4, finding['owners_model_b']).border = self.border
                ws.cell(row, 5, finding['rent_roll_c']).border = self.border
                ws.cell(row, 6, finding['stacking_plan_d']).border = self.border
                ws.cell(row, 7, finding['variance_a_b']).border = self.border
                ws.cell(row, 8, finding['variance_a_c']).border = self.border
                ws.cell(row, 9, finding['variance_a_d']).border = self.border
                
                # Highlight variances
                if finding['variance_a_b']:
                    ws.cell(row, 3).fill = self.yellow_fill
                    ws.cell(row, 4).fill = self.yellow_fill
                if finding['variance_a_c']:
                    ws.cell(row, 3).fill = self.yellow_fill
                    ws.cell(row, 5).fill = self.yellow_fill
                if finding['variance_a_d']:
                    ws.cell(row, 3).fill = self.yellow_fill
                    ws.cell(row, 6).fill = self.yellow_fill
                
                row += 1
        
        return row + 2
    
    def create_section_5_base_rent(self, findings: list, start_row: int):
        """Section 5: Base Rent with DETAILED escalation schedules"""
        
        logger.info("Creating Section 5: Base Rent & Escalations...")
        
        ws = self.wb["Audit Report"]
        row = start_row
        
        ws[f'A{row}'] = "5. For the following Tenants, there are discrepancies in Base Rent ($Amt/Yr)/ Base Rent Steps:"
        ws[f'A{row}'].font = self.bold_font
        row += 2
        
        # Column headers
        headers = ['Suite No.', 'Tenant Name', 'Prior NKF', '', "Owner's Model", '', 'Rent Roll', '', 'Leases', '']
        sub_headers = ['', '', 'Date', '$Amt/Yr', 'Date', '$Amt/Yr', 'Date', '$Amt/Yr', 'Date', '$Amt/Yr']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header)
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        row += 1
        
        for col_idx, sub_header in enumerate(sub_headers, 1):
            cell = ws.cell(row, col_idx, sub_header)
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        row += 1
        
        if not findings:
            ws.cell(row, 1, "No discrepancies found")
        else:
            for finding in findings:
                # Parse escalation schedules from JSON
                prior_schedule = self._parse_schedule(finding.get('prior_nkf', '[]'))
                owner_schedule = self._parse_schedule(finding.get('owners_model', '[]'))
                rentroll_schedule = self._parse_schedule(finding.get('rent_roll', '[]'))
                
                max_rows = max(len(prior_schedule), len(owner_schedule), len(rentroll_schedule))
                
                start_row = row
                
                # Write suite and tenant (merge vertically if multiple rows)
                if max_rows > 1:
                    ws.merge_cells(f'A{row}:A{row+max_rows-1}')
                    ws.merge_cells(f'B{row}:B{row+max_rows-1}')
                
                ws.cell(row, 1, finding['suite_number']).border = self.border
                ws.cell(row, 2, finding['tenant_name']).border = self.border
                
                # Write escalation rows
                for i in range(max_rows):
                    current_row = start_row + i
                    
                    # Prior NKF
                    if i < len(prior_schedule):
                        ws.cell(current_row, 3, prior_schedule[i].get('date', 'Base Rent')).border = self.border
                        ws.cell(current_row, 4, prior_schedule[i].get('amount_per_year', '')).border = self.border
                    else:
                        ws.cell(current_row, 3, '').border = self.border
                        ws.cell(current_row, 4, '').border = self.border
                    
                    # Owner's Model
                    if i < len(owner_schedule):
                        ws.cell(current_row, 5, owner_schedule[i].get('date', 'Base Rent')).border = self.border
                        ws.cell(current_row, 6, owner_schedule[i].get('amount_per_year', '')).border = self.border
                    else:
                        ws.cell(current_row, 5, '').border = self.border
                        ws.cell(current_row, 6, '').border = self.border
                    
                    # Rent Roll
                    if i < len(rentroll_schedule):
                        ws.cell(current_row, 7, rentroll_schedule[i].get('date', 'Base Rent')).border = self.border
                        ws.cell(current_row, 8, rentroll_schedule[i].get('amount_per_year', '')).border = self.border
                    else:
                        ws.cell(current_row, 7, '').border = self.border
                        ws.cell(current_row, 8, '').border = self.border
                    
                    # Leases
                    ws.cell(current_row, 9, 'NAV').border = self.border
                    ws.cell(current_row, 10, '').border = self.border
                    
                    # Highlight mismatches
                    if i == 0:  # Base rent row
                        prior_amt = prior_schedule[0].get('amount_per_year', '') if len(prior_schedule) > 0 else ''
                        owner_amt = owner_schedule[0].get('amount_per_year', '') if len(owner_schedule) > 0 else ''
                        
                        if prior_amt and owner_amt and prior_amt != owner_amt:
                            ws.cell(current_row, 4).fill = self.yellow_fill
                            ws.cell(current_row, 6).fill = self.yellow_fill
                
                row = start_row + max_rows
        
        return row + 2
    
    def _parse_schedule(self, schedule_json: str) -> List[Dict]:
        """Parse escalation schedule from JSON string"""
        try:
            return json.loads(schedule_json)
        except:
            return []
    
    def create_section_6_rent_per_sf(self, findings: list, start_row: int):
        """Section 6: Rent Per SF"""
        
        logger.info("Creating Section 6: Rent Per SF...")
        
        ws = self.wb["Audit Report"]
        row = start_row
        
        ws[f'A{row}'] = "6. For the following Tenants, there are discrepancies in Rent Per SF:"
        ws[f'A{row}'].font = self.bold_font
        row += 2
        
        headers = ['Suite No.', 'Tenant Name', 'Prior NKF', "Owner's Model", 
                  'Rent Roll', 'Variance (A-B)', 'Variance (A-C)']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header)
            cell.font = self.bold_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        row += 1
        
        if not findings:
            ws.cell(row, 1, "No discrepancies found")
        else:
            for finding in findings:
                ws.cell(row, 1, finding['suite_number']).border = self.border
                ws.cell(row, 2, finding['tenant_name']).border = self.border
                ws.cell(row, 3, finding['prior_rate']).border = self.border
                ws.cell(row, 4, finding['owner_rate']).border = self.border
                ws.cell(row, 5, finding['rentroll_rate']).border = self.border
                ws.cell(row, 6, finding['variance_a_b']).border = self.border
                ws.cell(row, 7, finding['variance_a_c']).border = self.border
                
                if finding['variance_a_b']:
                    ws.cell(row, 3).fill = self.yellow_fill
                    ws.cell(row, 4).fill = self.yellow_fill
                if finding['variance_a_c']:
                    ws.cell(row, 3).fill = self.yellow_fill
                    ws.cell(row, 5).fill = self.yellow_fill
                
                row += 1
        
        return row + 2
    
    def set_column_widths(self):
        """Set column widths"""
        ws = self.wb["Audit Report"]
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
    
    def generate_report(self, findings: dict) -> str:
        """Generate complete report"""
        
        logger.info("\n" + "="*80)
        logger.info("GENERATING IMPROVED AUDIT REPORT")
        logger.info("="*80)
        
        row = self.create_section_1_lease_status(findings.get('lease_status', []))
        row = self.create_section_2_tenant_name(findings.get('tenant_name', []), row)
        row = self.create_section_3_lease_expiration(findings.get('lease_expiration', []), row)
        row = self.create_section_4_leased_area(findings.get('leased_area', []), row)
        row = self.create_section_5_base_rent(findings.get('base_rent', []), row)
        row = self.create_section_6_rent_per_sf(findings.get('rent_per_sf', []), row)
        
        self.set_column_widths()
        
        self.wb.save(self.output_filename)
        logger.info(f"\n Report saved: {self.output_filename}")
        logger.info("="*80)
        
        return self.output_filename


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    
    input_dir = Path(args.input)
    
    # Load findings
    findings = {}
    for category in ['lease_status', 'tenant_name', 'lease_expiration', 
                     'leased_area', 'base_rent', 'rent_per_sf']:
        csv_file = input_dir / f"{category}_findings.csv"
        if csv_file.exists():
            findings[category] = pd.read_csv(csv_file).to_dict('records')
            logger.info(f"Loaded: {category} ({len(findings[category])} findings)")
        else:
            findings[category] = []
    
    # Generate report
    reporter = ImprovedAuditReporter(args.output)
    report_path = reporter.generate_report(findings)
    
    logger.info(f"\n Report complete: {report_path}")


if __name__ == "__main__":
    main()